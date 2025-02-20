# -*- coding: utf-8 -*-

import openpyxl
import yaml
import json
from configparser import ConfigParser
from common.logger import logger
from enum import Enum
import os

BASE_PATH = os.path.abspath(__file__).split(
    "PremomApiTest")[0] + "PremomApiTest" + ""


class CaseEnum(Enum):
    API_EXEC = 1
    CASE_ID = 2
    CASE_FEATURE = 3
    CASE_TITLE = 4
    API_PATH = 5
    API_HEADER = 6
    API_METHOD = 7
    API_PK = 8
    API_DATA = 9
    API_FILE = 10
    API_EXTRACT = 11
    API_EXPECTED = 12


class ReadExcel:

    def __init__(self, filename):
        self.filename = filename
        self.workbook = openpyxl.load_workbook(self.filename)
        self.sheets = self.workbook.sheetnames

    def read_excel(self, sheet):
        case_data = []
        if sheet == "":
            sheets = self.sheets
        else:
            sheets = [sheet]
        logger.info(sheet)
        for sheet in sheets:
            wb = self.workbook[sheet]
            logger.info(wb.cell(row=1, column=1).value)
            logger.info(wb.max_column)
            max_row = self.workbook[sheet].max_row
            for i in range(2, max_row + 1):
                _dict = {}
                if wb.cell(row=i,
                           column=CaseEnum.API_EXEC.value).value == 'Yes':
                    _dict["id"] = wb.cell(row=i,
                                          column=CaseEnum.CASE_ID.value).value
                    _dict["feature"] = wb.cell(
                        row=i, column=CaseEnum.CASE_FEATURE.value).value
                    _dict["title"] = wb.cell(
                        row=i, column=CaseEnum.CASE_TITLE.value).value
                    _dict["url"] = wb.cell(
                        row=i, column=CaseEnum.API_PATH.value).value
                    _dict["header"] = wb.cell(
                        row=i, column=CaseEnum.API_HEADER.value).value
                    _dict["method"] = wb.cell(
                        row=i, column=CaseEnum.API_METHOD.value).value
                    _dict["pk"] = wb.cell(row=i,
                                          column=CaseEnum.API_PK.value).value
                    _dict["data"] = wb.cell(
                        row=i, column=CaseEnum.API_DATA.value).value
                    _dict["file"] = wb.cell(
                        row=i, column=CaseEnum.API_FILE.value).value
                    _dict["extract"] = wb.cell(
                        row=i, column=CaseEnum.API_EXTRACT.value).value
                    _dict["validate"] = wb.cell(
                        row=i, column=CaseEnum.API_EXPECTED.value).value

                    case_data.append(_dict)

        return case_data


class MyConfigParser(ConfigParser):

    def __init__(self, defaults=None):
        ConfigParser.__init__(self, defaults=defaults)

    def optionxform(self, optionstr):
        return optionstr


class ReadFileData():

    def __init__(self):
        pass

    def load_yaml(self, file_path):
        logger.info("Loading {} file...".format(file_path))
        with open(file_path, encoding='utf-8') as f:
            data = yaml.safe_load(f)
        logger.info("Data loaded: {}".format(data))
        return data

    def load_json(self, file_path):
        logger.info("Loading {} file...".format(file_path))
        with open(file_path, encoding='utf-8') as f:
            data = json.load(f)
        logger.info("Data loaded: {}".format(data))
        return data

    def load_ini(self, file_path):
        logger.info("Loading {} file...".format(file_path))
        config = MyConfigParser()
        config.read(file_path, encoding="UTF-8")
        data = dict(config._sections)
        return data


data = ReadFileData()

# if __name__ == "__main__":
#     filePath = BASE_PATH + "/data/AutoPytestExecl.xlsx"
#
#     # 创建 ReadExcel 类的实例
#     excel_reader = ReadExcel(filePath)
#
#     # 调用 read_excel 方法来读取指定工作表的数据
#     data = excel_reader.read_excel("ceshi")
#
#     # 现在 data 中包含了从 "ceshi" 工作表读取到的数据
#     print(data)
